#!/usr/bin/env python
"""Inserta firmas de operadores en Planilla_CLARO_FUTBOLFEST.xlsx.

Lee las firmas PNG de una carpeta local (descargadas del VPS) y las inserta
en la columna FIRMA del Excel, una por operador, identificando por cédula.

PROTEGE el archivo original:
  - Hace un backup ``.bak`` antes de guardar.
  - NO reordena filas, NO reescribe celdas, NO cambia estilos.
  - Solo añade imágenes encima de las celdas de FIRMA.

Estructura detectada (todas las hojas):
  - Encabezados: fila 8
  - CÉDULA: columna E (5)
  - FIRMA:  columna L (12)
  - Datos:  filas 9..28

Uso:
    python backend/scripts/stamp_signatures.py [carpeta_firmas] [ruta_excel]

Ejemplo:
    python backend/scripts/stamp_signatures.py "C:\\Users\\Karen\\Downloads\\firmas_evento"
"""
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# --- Configuración (hardcodeada de la inspección) --------------------------
DEFAULT_XLSX = Path(r"C:\Users\Karen\Downloads\Planilla_CLARO_FUTBOLFEST.xlsx")
DEFAULT_FIRMAS = Path(r"C:\Users\Karen\Downloads\firmas_evento")

HEADER_ROW = 8
COL_CEDULA = 5    # E
COL_FIRMA = 12    # L
FIRST_DATA_ROW = 9
LAST_DATA_ROW = 28

# Margen interno (fracción) para que la firma no toque los bordes de la celda
SCALE_MARGIN = 0.90
# ---------------------------------------------------------------------------


def _normalize_cedula(raw) -> str:
    """Deja solo dígitos."""
    if raw is None:
        return ""
    return "".join(re.findall(r"\d+", str(raw)))


def _load_signatures(firmas_dir: Path) -> dict[str, Path]:
    """Carga {cedula: ruta_png} de la carpeta de firmas.

    Busca recursivamente (rglob) para tolerar carpetas anidadas tipo
    ``firmas_evento/firmas_evento/*.png`` (común al copiar con scp/docker cp).
    """
    mapping: dict[str, Path] = {}
    if not firmas_dir.exists():
        return mapping
    for png in firmas_dir.rglob("*.png"):
        ced = _normalize_cedula(png.stem)
        if ced:
            mapping[ced] = png
    return mapping


def _cell_size_px(ws, row: int, col: int) -> tuple[float, float]:
    """Calcula (ancho_px, alto_px) de una celda.

    - Ancho de columna: Excel guarda "characters"; convertir a píxeles.
      Fórmula: pixels = round(width * 7 + 5)  (aprox Calibri 11).
    - Alto de fila: se guarda en puntos; convertir a píxeles (1 pt = 4/3 px).
    """
    letra = get_column_letter(col)
    col_dim = ws.column_dimensions.get(letra)
    width_chars = col_dim.width if col_dim and col_dim.width else 8.43  # default
    width_px = round(width_chars * 7 + 5)

    row_dim = ws.row_dimensions.get(row)
    height_pts = row_dim.height if row_dim and row_dim.height else 15.0  # default
    height_px = height_pts * (4 / 3)

    return width_px, height_px


def _scaled_image(src_png: Path, max_w_px: float, max_h_px: float) -> XlImage | None:
    """Crea una openpyxl.Image escalada para caber en (max_w, max_h) px.

    Mantiene la proporción original. Aplica SCALE_MARGIN para dejar aire.
    """
    try:
        pil = PILImage.open(src_png)
        ow, oh = pil.size
        pil.close()
    except Exception as exc:
        print(f"   ⚠️  No se pudo leer imagen {src_png.name}: {exc}")
        return None

    if not ow or not oh:
        return None

    avail_w = max_w_px * SCALE_MARGIN
    avail_h = max_h_px * SCALE_MARGIN
    scale = min(avail_w / ow, avail_h / oh)
    new_w = max(1, int(round(ow * scale)))
    new_h = max(1, int(round(oh * scale)))

    img = XlImage(str(src_png))
    img.width = new_w
    img.height = new_h
    return img


def stamp_workbook(xlsx_path: Path, firmas_dir: Path) -> None:
    print(f"\n{'=' * 70}")
    print(f"🖋️  Insertando firmas en: {xlsx_path}")
    print(f"📁 Carpeta de firmas: {firmas_dir}")

    if not xlsx_path.exists():
        print(f"   ❌ Excel no encontrado: {xlsx_path}")
        return
    if not firmas_dir.exists():
        print(f"   ❌ Carpeta de firmas no encontrada: {firmas_dir}")
        print(f"      Descarga las firmas del VPS primero (ver Fase 0).")
        return

    firmas = _load_signatures(firmas_dir)
    print(f"   ✅ {len(firmas)} firmas cargadas de la carpeta.")
    if not firmas:
        print("   ⚠️  No hay PNG en la carpeta. Nada que insertar.")
        return

    # --- Backup .bak ---
    bak = xlsx_path.with_suffix(xlsx_path.suffix + ".bak")
    shutil.copy2(xlsx_path, bak)
    print(f"   💾 Backup creado: {bak.name}")

    # --- Cargar Excel ---
    wb = openpyxl.load_workbook(xlsx_path)

    total_inserted = 0
    total_missing = 0
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Hoja: '{sheet_name}' ---")

        inserted = 0
        missing: list[tuple[int, str]] = []

        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            ced_raw = ws.cell(row=row, column=COL_CEDULA).value
            cedula = _normalize_cedula(ced_raw)
            if not cedula:
                continue
            total_rows += 1

            png_path = firmas.get(cedula)
            if not png_path:
                missing.append((row, cedula))
                total_missing += 1
                continue

            # Tamaño de la celda de firma
            w_px, h_px = _cell_size_px(ws, row, COL_FIRMA)
            img = _scaled_image(png_path, w_px, h_px)
            if img is None:
                missing.append((row, cedula))
                total_missing += 1
                continue

            cell_ref = f"{get_column_letter(COL_FIRMA)}{row}"
            ws.add_image(img, cell_ref)
            inserted += 1
            total_inserted += 1

        print(f"   ✅ Firmas insertadas: {inserted}")
        if missing:
            print(f"   ⚠️  Sin firma ({len(missing)}): {', '.join(c for _, c in missing)}")

    # --- Guardar (mismo archivo) ---
    wb.save(xlsx_path)
    wb.close()

    print(f"\n{'=' * 70}")
    print(f"✅ Proceso completado.")
    print(f"   Filas con cédula: {total_rows}")
    print(f"   Firmas insertadas: {total_inserted}")
    print(f"   Sin firma en carpeta: {total_missing}")
    print(f"   Archivo guardado: {xlsx_path}")
    print(f"   Backup: {bak.name}")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    firmas = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FIRMAS
    xlsx = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_XLSX
    stamp_workbook(xlsx, firmas)