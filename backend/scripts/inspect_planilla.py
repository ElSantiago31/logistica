#!/usr/bin/env python
"""Inspección de Planilla_CLARO_FUTBOLFEST.xlsx.

Reporta la estructura de cada hoja para detectar dónde está la cédula y la
firma. NO modifica el archivo.

Uso:
    python backend/scripts/inspect_planilla.py [ruta_al_xlsx]
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# Ruta por defecto: la planilla del usuario en Downloads
DEFAULT_PATH = Path(r"C:\Users\Karen\Downloads\Planilla_CLARO_FUTBOLFEST.xlsx")

# Palabras clave para detectar columnas
KEYWORDS_CEDULA = ["cedula", "cédula", "cc", "doc", "identificacion", "identificación", "no."]
KEYWORDS_FIRMA = ["firma", "autografa", "autógrafa", "sign", "firms"]


def _norm(text: str) -> str:
    """Normaliza texto: minúsculas, sin acentos, sin espacios extra."""
    import unicodedata
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()


def _normalize_cedula(raw) -> str:
    """Deja solo dígitos de una cédula."""
    if raw is None:
        return ""
    return "".join(re.findall(r"\d+", str(raw)))


def _detect_header_row(ws, max_scan: int = 15) -> tuple[int | None, dict[int, str]]:
    """Escanea las primeras ``max_scan`` filas buscando la fila de encabezados.

    Retorna (fila_header, {col_idx: texto_header}).
    """
    best_row = None
    best_headers: dict[int, str] = {}
    best_score = 0

    for row in range(1, min(max_scan + 1, ws.max_row + 1)):
        headers: dict[int, str] = {}
        score = 0
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip():
                headers[col] = _norm(val)
                # Puntúa filas que parezcan encabezados
                norm = _norm(val)
                if any(k in norm for k in KEYWORDS_CEDULA):
                    score += 2
                if any(k in norm for k in KEYWORDS_FIRMA):
                    score += 2
                if norm in ("nombres", "apellidos", "nombre", "nombre completo"):
                    score += 1
                if norm in ("no", "n°", "item", "#", "n"):
                    score += 1
        if score > best_score:
            best_score = score
            best_row = row
            best_headers = headers

    return (best_row, best_headers) if best_row else (None, {})


def _find_col(headers: dict[int, str], keywords: list[str]) -> int | None:
    """Busca la columna cuyo header contiene alguna keyword."""
    for col, txt in headers.items():
        for kw in keywords:
            if kw in txt:
                return col
    return None


def inspect_workbook(path: Path) -> None:
    print(f"\n{'=' * 70}")
    print(f"📄 Inspeccionando: {path}")
    print(f"   Existe: {path.exists()}")
    if not path.exists():
        print("   ❌ Archivo no encontrado.")
        return

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    print(f"   Hojas ({len(wb.sheetnames)}): {wb.sheetnames}")
    print("=" * 70)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Hoja: '{sheet_name}' ---")
        print(f"   Dimensiones: {ws.max_row} filas × {ws.max_column} columnas")

        # Detectar fila de encabezados
        header_row, headers = _detect_header_row(ws)
        if header_row:
            print(f"   🔹 Fila de encabezados detectada: fila {header_row}")
            print("   Encabezados:")
            for col, txt in sorted(headers.items()):
                letra = get_column_letter(col)
                print(f"      {letra}{header_row} (col {col}): {txt!r}")
        else:
            print("   ⚠️  No se detectó fila de encabezados clara.")
            print("   Primeras 5 filas (raw):")
            for r in range(1, min(6, ws.max_row + 1)):
                vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
                print(f"      fila {r}: {vals}")
            continue

        # Detectar columna de cédula y firma
        col_ced = _find_col(headers, KEYWORDS_CEDULA)
        col_firma = _find_col(headers, KEYWORDS_FIRMA)
        print(f"   📍 Columna CÉDULA detectada: {col_ced} ({get_column_letter(col_ced) if col_ced else '?'})")
        print(f"   📍 Columna FIRMA detectada: {col_firma} ({get_column_letter(col_firma) if col_firma else '?'})")

        if col_ced is None:
            print("   ⚠️  No se detectó columna de cédula. Revisa los encabezados arriba.")
            continue

        # Contar cédulas (filas con datos a partir de header_row+1)
        cedulas = []
        for r in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(row=r, column=col_ced).value
            ced = _normalize_cedula(raw)
            if ced:
                cedulas.append((r, ced))

        print(f"   👥 Cédulas encontradas: {len(cedulas)}")
        if cedulas:
            print(f"      Primera: fila {cedulas[0][0]} → {cedulas[0][1]}")
            print(f"      Última:  fila {cedulas[-1][0]} → {cedulas[-1][1]}")

        # Reportar ancho de columna de firma (para escalar imágenes)
        if col_firma:
            letra = get_column_letter(col_firma)
            col_dim = ws.column_dimensions.get(letra)
            width = col_dim.width if col_dim and col_dim.width else None
            print(f"   📐 Ancho columna firma ({letra}): {width}")

        # Alto de filas de datos (muestra)
        if cedulas:
            r0 = cedulas[0][0]
            rd = ws.row_dimensions.get(r0)
            h = rd.height if rd and rd.height else None
            print(f"   📐 Alto fila datos (ej. fila {r0}): {h}")

    print("\n" + "=" * 70)
    print("✅ Inspección completa.\n")


if __name__ == "__main__":
    p = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PATH
    inspect_workbook(p)