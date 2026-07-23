"""Servicio para generar el Excel de planilla de pago.

Toma la plantilla ``Planilla_Logistica_Eventos.xlsx`` (con logo, formato y
encabezado) y genera un workbook con:

- ``group_by="coordinator"`` (default): una hoja por cada coordinador.
- ``group_by="none"``: una hoja plana (lista única) titulada con el evento.

Dentro de cada hoja, los operadores se ordenan por:

- ``sort_by="lastname"`` (default): por apellido (alfabético).
- ``sort_by="document"``: por número de cédula (numérico cuando es posible).

Si una hoja tiene más de 20 operadores, se generan hojas adicionales
paginadas (``COORDINADOR (1)``, ``COORDINADOR (2)``, ... o
``EVENTO (1)``, ``EVENTO (2)`` ...).

Solo se incluyen los operadores con ``status='checked_in'``.

Estructura de la plantilla (1 hoja "Planilla"):
    Encabezado:
        D5  → COORDINADOR GENERAL  (celda combinada D5:J5)
        D6  → NOMBRE DEL EVENTO    (celda combinada D6:G6)
        D7  → FECHA                (celda combinada D7:H7)
        K7  → LUGAR                (celda combinada K7:M7)
    Tabla (filas 9-28, 20 operadores pre-numerados):
        B=No | C=NOMBRES | D=APELLIDOS | E=CEDULA | F=DIRECCION
        G=CELULAR | H=COORDINADOR | I=No CHAQ | J=No GORRA
        K=VALOR | L=FIRMA | M=No DSE   (estos quedan vacíos)
"""
import io
import re
import copy as _copy
import base64
import logging
import unicodedata
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import (
    OneCellAnchor,
    AnchorMarker,
    XDRPositiveSize2D,
)

logger = logging.getLogger(__name__)

# --- Colores para resaltar filas (novedades / vetos) ---
# Tonos claros para mantener legibilidad al imprimir en blanco y negro.
# Rojo claro (vetado) tiene prioridad sobre amarillo claro (novedad).
BAN_FILL = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")  # red-200
INCIDENT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber-100
# Última columna de datos en la plantilla (M=13). Se usa para pintar toda la fila.
LAST_COL = 13

# --- Configuración de la plantilla ---
TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "templates"
    / "Planilla_Logistica_Eventos.xlsx"
)
TEMPLATE_SHEET = "Planilla"
ROWS_PER_PAGE = 20
FIRST_DATA_ROW = 9  # primera fila de datos en la plantilla
LAST_DATA_ROW = FIRST_DATA_ROW + ROWS_PER_PAGE - 1  # 28

# --- Columnas (1-indexed) según el header de la plantilla ---
COL_NO = 2           # B  — No (viene pre-numerado en la plantilla)
COL_NOMBRES = 3      # C
COL_APELLIDOS = 4    # D
COL_CEDULA = 5       # E
COL_DIRECCION = 6    # F
COL_CELULAR = 7      # G
COL_COORDINADOR = 8  # H
COL_CHAQ = 9         # I
COL_GORRA = 10       # J
COL_FIRMA = 12       # L — columna donde van las firmas embebidas
# K=VALOR, M=No DSE → se dejan vacíos (no tenemos el dato)

# --- Celdas del encabezado (top-left de la celda combinada) ---
CELL_COORD = "D5"
CELL_EVENTO = "D6"
CELL_FECHA = "D7"
CELL_LUGAR = "K7"


def _split_name(full_name: str) -> tuple[str, str]:
    """Divide un nombre completo en (nombres, apellidos).

    El primer token se considera nombre y el resto apellidos.
    Ej: "JUAN CARLOS PEREZ GOMEZ" → ("JUAN CARLOS", "PEREZ GOMEZ")
    Pero el usuario pidió: primer token = nombre, resto = apellidos.
    """
    if not full_name:
        return ("", "")
    parts = full_name.strip().split(None, 1)
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], parts[1])


def _norm_sort_text(text: str) -> str:
    """Normaliza texto para ordenamiento: quita acentos y pasa a mayúsculas.

    Ej: "Pérez Gómez" → "PEREZ GOMEZ"
    """
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    return nfkd.encode("ascii", "ignore").decode("ascii").upper().strip()


def _sort_key_by_lastname(full_name: str) -> tuple[str, str]:
    """Clave de ordenamiento por APELLIDO (case-insensitive, sin acentos).

    Usa la MISMA lógica que ``_split_name`` para extraer el apellido:
    todo después del primer token se considera apellido.

    Ej: "JUAN CARLOS PEREZ GOMEZ" → clave ("PEREZ GOMEZ", "JUAN CARLOS")

    Operadores sin apellido (un solo token) se ordenan por ese token como
    apellido.

    Returns:
        Tuple (apellido_norm, nombre_norm) para usar en ``sorted(key=...)``.
    """
    nombres, apellidos = _split_name(full_name)

    # Si no hay apellido (un solo token), usar el nombre como apellido
    # para que se ubique alfabéticamente en lugar de quedar agrupado al final.
    return (
        _norm_sort_text(apellidos) or _norm_sort_text(nombres),
        _norm_sort_text(nombres),
    )


def _sort_key_by_document(op: dict) -> tuple:
    """Clave de ordenamiento por número de cédula (numérico cuando es posible).

    Extrae la parte numérica inicial del ``document_number`` para ordenar
    correctamente (``"1020"`` < ``"52000"``). Si no es numérico, ordena como
    string al final de la lista numérica.

    Returns:
        Tuple ``(is_numeric, numeric_value, raw)`` para usar en ``sorted``.
    """
    raw = (op.get("document_number") or "").strip()
    if not raw:
        return (False, 0, "")
    # Extraer dígitos iniciales
    m = re.match(r"^(\d+)", raw)
    if m:
        return (True, int(m.group(1)), raw)
    return (False, 0, raw.upper())


# Caracteres prohibidos en nombres de hojas de Excel
_SHEET_INVALID_CHARS = re.compile(r"[\[\]\*\/\\\?\:]")


def _sanitize_sheet_title(title: str, page_num: int | None = None) -> str:
    """Sanea un título para usarlo como nombre de hoja Excel.

    - Elimina caracteres prohibidos: ``[ ] * / \\ ? :``
    - Recorta a 31 caracteres (límite de Excel)
    - Si ``page_num`` no es None, añade `` (N)`` reservando espacio
    """
    cleaned = _SHEET_INVALID_CHARS.sub(" ", title or "").strip()
    if page_num is not None:
        suffix = f" ({page_num})"
        # Reservar espacio para el sufijo
        max_len = 31 - len(suffix)
        cleaned = cleaned[:max_len] + suffix
    else:
        cleaned = cleaned[:31]
    return cleaned or "Hoja"


def _fmt_date(dt: datetime | None) -> str:
    """Formatea una fecha como DD/MM/YYYY para la planilla."""
    if not dt:
        return ""
    # Si tiene tzinfo, convertir a Bogotá
    try:
        from app.routers.payroll import BOGOTA_TZ
        if dt.tzinfo is not None:
            dt = dt.astimezone(BOGOTA_TZ)
    except Exception:
        pass
    return dt.strftime("%d/%m/%Y")


# --- EMU (English Metric Units) para posicionamiento de imágenes ---
# 1 píxel (a 96 DPI) = 9525 EMU. Se usa para anclar imágenes con precisión.
_EMU_PER_PX = 9525
# Alto de fila (en puntos) cuando la fila lleva una firma. Permite que la
# firma sea legible y no quede comprimida.
_SIGNATURE_ROW_HEIGHT_PT = 50


def _decode_signature(signature_data: str | None) -> bytes | None:
    """Decodifica un base64 PNG/JPG de firma a bytes crudos.

    Maneja el prefijo ``data:image/png;base64,`` (firmas desde el navegador).
    Reutiliza el patrón de ``app.services.invoice_pdf._decode_signature``.

    Returns:
        Bytes de la imagen, o ``None`` si la entrada es vacía/inválida.
    """
    if not signature_data:
        return None
    try:
        raw = signature_data.strip()
        if "," in raw and raw.startswith("data:image"):
            raw = raw.split(",", 1)[1]
        return base64.b64decode(raw)
    except Exception as exc:
        logger.warning("No se pudo decodificar la firma (base64): %s", exc)
        return None


def _get_image_size(img_bytes: bytes) -> tuple[int, int]:
    """Devuelve ``(width_px, height_px)`` de una imagen usando PIL.

    Returns:
        Tupla con dimensiones en píxeles, o ``(0, 0)`` si no se puede leer.
    """
    try:
        from PIL import Image as PILImage
        pil = PILImage.open(io.BytesIO(img_bytes))
        return (pil.width, pil.height)
    except Exception as exc:
        logger.warning("No se pudo leer el tamaño de la firma: %s", exc)
        return (0, 0)


def _col_letter(col: int) -> str:
    """Convierte índice de columna (1-indexed) a letra (1→A, 12→L)."""
    return openpyxl.utils.get_column_letter(col)


def _add_centered_signature(
    ws,
    row: int,
    col: int,
    signature_data: str,
    fill_ratio: float = 0.8,
) -> bool:
    """Embebe una firma centrada en la celda ``(row, col)``.

    La firma se escala manteniendo el aspect ratio para que ocupe ``fill_ratio``
    (80% por defecto) de la celda, de modo que se vean los bordes y no quede
    pegada a los límites. Se centra horizontal y verticalmente usando un
    ``OneCellAnchor`` con offsets en EMU.

    Fórmulas de conversión (Excel → px → EMU):
        - Ancho de columna: ``px = round(width_chars * 7) + 5``
        - Alto de fila: ``px = round(height_pt * 96 / 72)``
        - 1 px = 9525 EMU

    Args:
        ws: Worksheet de openpyxl.
        row: Fila (1-indexed) de la celda destino.
        col: Columna (1-indexed) de la celda destino.
        signature_data: Firma en base64 PNG/JPG (con o sin prefijo data:).
        fill_ratio: Fracción de la celda que ocupa la firma (0-1). Default 0.8.

    Returns:
        ``True`` si se embebió correctamente, ``False`` si la firma era
        inválida o falló el posicionamiento.
    """
    sig_bytes = _decode_signature(signature_data)
    if not sig_bytes:
        return False

    img_w, img_h = _get_image_size(sig_bytes)
    if img_w <= 0 or img_h <= 0:
        return False

    # --- Dimensiones de la celda en píxeles ---
    col_dim = ws.column_dimensions.get(_col_letter(col))
    col_width_chars = col_dim.width if (col_dim and col_dim.width) else 8.43
    col_w_px = round(col_width_chars * 7) + 5

    row_dim = ws.row_dimensions.get(row)
    row_height_pt = row_dim.height if (row_dim and row_dim.height) else 15
    row_h_px = round(row_height_pt * 96 / 72)

    # --- Tamaño objetivo (fill_ratio % de la celda) ---
    target_w_px = col_w_px * fill_ratio
    target_h_px = row_h_px * fill_ratio

    # --- Escalar manteniendo aspect ratio ---
    scale = min(target_w_px / img_w, target_h_px / img_h)
    final_w_px = img_w * scale
    final_h_px = img_h * scale

    # --- Offsets de centrado (en px → EMU) ---
    x_off_px = (col_w_px - final_w_px) / 2
    y_off_px = (row_h_px - final_h_px) / 2
    x_off_emu = int(max(0, x_off_px) * _EMU_PER_PX)
    y_off_emu = int(max(0, y_off_px) * _EMU_PER_PX)

    # --- Tamaño final en EMU ---
    ext_w_emu = int(final_w_px * _EMU_PER_PX)
    ext_h_emu = int(final_h_px * _EMU_PER_PX)

    # --- Crear y anclar la imagen ---
    try:
        img = XlImage(io.BytesIO(sig_bytes))
        # OneCellAnchor: la imagen parte de una celda con offsets fijos y
        # tiene un tamaño absoluto (no se estira si la celda cambia).
        marker = AnchorMarker(col=col - 1, colOff=x_off_emu, row=row - 1, rowOff=y_off_emu)
        anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(cx=ext_w_emu, cy=ext_h_emu),
        )
        img.anchor = anchor
        ws.add_image(img)
        return True
    except Exception as exc:
        logger.warning("No se pudo embeber la firma en celda (%s,%s): %s", row, col, exc)
        return False


def _copy_sheet_with_images(wb: openpyxl.Workbook, src_sheet: openpyxl.worksheet.worksheet.Worksheet, new_title: str):
    """Copia una hoja dentro del workbook preservando imágenes (logo).

    ``openpyxl.Workbook.copy_worksheet`` NO copia imágenes, así que las
    re-agregamos manualmente tras la copia.

    Nota: Se usa ``deepcopy`` en vez de ``copy`` para evitar que múltiples
    hojas compartan el mismo ``BytesIO`` de la imagen, lo cual causa
    ``ValueError: I/O operation on closed file`` al serializar.
    """
    new_ws = wb.copy_worksheet(src_sheet)
    new_ws.title = new_title

    # Re-agregar imágenes (logo) — copy_worksheet no las copia
    if src_sheet._images:
        for img in src_sheet._images:
            new_img = _copy.deepcopy(img)
            new_ws._images.append(new_img)

    return new_ws


def _fill_header(ws, *, coordinator: str, event_name: str, event_date, event_location: str):
    """Rellena las celdas del encabezado de una hoja."""
    ws[CELL_COORD] = coordinator or ""
    ws[CELL_EVENTO] = event_name or ""
    ws[CELL_FECHA] = _fmt_date(event_date)
    ws[CELL_LUGAR] = event_location or ""


def _fill_operators(ws, operators: list[dict], with_signatures: bool = False):
    """Rellena las filas de operadores (a partir de la fila 9).

    ``operators`` es una lista de dicts con las claves:
        full_name, document_number, address, phone,
        coordinator_name, jacket_number, cap_number,
        is_banned (bool), has_incident (bool),
        signature_data (str|None, base64 PNG — solo cuando with_signatures)

    Resalta toda la fila de color según el estado del operador:
        - **Rojo claro** si tiene veto activo (``is_banned=True``).
        - **Amarillo claro** si tiene alguna novedad en el evento
          (``has_incident=True``).
        El veto (rojo) tiene prioridad visual sobre la novedad (amarillo).

    Cuando ``with_signatures=True`` y el operador tiene ``signature_data``,
    se aumenta la altura de la fila a ``_SIGNATURE_ROW_HEIGHT_PT`` (50pt) y
    se embebe la firma centrada en la columna L (``COL_FIRMA``).
    """
    for idx, op in enumerate(operators):
        row = FIRST_DATA_ROW + idx
        if row > LAST_DATA_ROW:
            break  # seguridad — el paginado ya fragmentó en bloques de 20
        # Usar first_name/last_name directamente (campos separados de la BD).
        # Fallback a _split_name(full_name) solo si no vienen (retrocompatibilidad).
        nombres = (op.get("first_name") or "").strip()
        apellidos = (op.get("last_name") or "").strip()
        if not nombres and not apellidos:
            nombres, apellidos = _split_name(op.get("full_name", ""))
        ws.cell(row=row, column=COL_NOMBRES, value=nombres)
        ws.cell(row=row, column=COL_APELLIDOS, value=apellidos)
        ws.cell(row=row, column=COL_CEDULA, value=op.get("document_number", ""))
        ws.cell(row=row, column=COL_DIRECCION, value=op.get("address", ""))
        ws.cell(row=row, column=COL_CELULAR, value=op.get("phone", ""))
        ws.cell(row=row, column=COL_COORDINADOR, value=op.get("coordinator_name", ""))
        ws.cell(row=row, column=COL_CHAQ, value=op.get("jacket_number", ""))
        ws.cell(row=row, column=COL_GORRA, value=op.get("cap_number", ""))

        # --- Embeber firma (opcional) en la columna L ---
        # Solo si with_signatures=True y el operador tiene signature_data.
        # Se aumenta la altura de la fila ANTES de calcular el centrado,
        # porque _add_centered_signature lee ws.row_dimensions[row].height.
        sig_data = op.get("signature_data") if with_signatures else None
        if sig_data:
            ws.row_dimensions[row].height = _SIGNATURE_ROW_HEIGHT_PT
            _add_centered_signature(ws, row, COL_FIRMA, sig_data)

        # --- Resaltar fila completa si el operador tiene novedad o veto ---
        # Prioridad: veto (rojo) > novedad (amarillo).
        fill = None
        if op.get("is_banned"):
            fill = BAN_FILL
        elif op.get("has_incident"):
            fill = INCIDENT_FILL

        if fill is not None:
            for col in range(COL_NO, LAST_COL + 1):
                ws.cell(row=row, column=col).fill = fill


def _sort_operators(operators: list[dict], sort_by: str) -> list[dict]:
    """Ordena una lista de operadores según el criterio indicado.

    Args:
        operators: Lista de dicts de operadores.
        sort_by: ``"lastname"`` (default) o ``"document"``.

    Returns:
        Nueva lista ordenada (no muta la original).
    """
    if sort_by == "document":
        return sorted(operators, key=_sort_key_by_document)
    # default: lastname — usar first_name/last_name si están disponibles
    # (campos separados de la BD). Si no, fallback a _split_name(full_name).
    def _lastname_key(op: dict) -> tuple[str, str]:
        last = (op.get("last_name") or "").strip()
        first = (op.get("first_name") or "").strip()
        if last or first:
            return (
                _norm_sort_text(last) or _norm_sort_text(first),
                _norm_sort_text(first),
            )
        # Fallback (retrocompatibilidad): usar full_name
        return _sort_key_by_lastname(op.get("full_name", ""))

    return sorted(operators, key=_lastname_key)


def _render_pages(
    *,
    template_wb: openpyxl.Workbook,
    template_ws,
    event_name: str,
    event_date: datetime | None,
    event_location: str,
    sheet_label: str,
    operators: list[dict],
    sort_by: str = "lastname",
    with_signatures: bool = False,
) -> list[str]:
    """Crea una o varias hojas (paginadas cada ROWS_PER_PAGE) para un grupo.

    Args:
        sheet_label: Nombre base para las hojas (p.ej. nombre del coordinador
            o el nombre del evento cuando ``group_by="none"``).
        operators: Operadores del grupo (se ordenan aquí según ``sort_by``).
        with_signatures: Si ``True``, embebe las firmas (clave ``signature_data``
            de cada operador) centradas en la columna L.

    Returns:
        Lista con los títulos de las hojas creadas.
    """
    operators = _sort_operators(list(operators), sort_by)
    sheets_created = []
    total_pages = max(1, (len(operators) + ROWS_PER_PAGE - 1) // ROWS_PER_PAGE)

    for page_num in range(total_pages):
        start = page_num * ROWS_PER_PAGE
        end = start + ROWS_PER_PAGE
        page_ops = operators[start:end]

        # Título de la hoja (con sufijo de página si hay más de una)
        page_suffix = page_num + 1 if total_pages > 1 else None
        sheet_title = _sanitize_sheet_title(sheet_label, page_num=page_suffix)

        new_ws = _copy_sheet_with_images(template_wb, template_ws, sheet_title)
        _fill_header(
            new_ws,
            coordinator="",  # "Coordinador General" (D5) se deja vacío
            event_name=event_name,
            event_date=event_date,
            event_location=event_location,
        )
        _fill_operators(new_ws, page_ops, with_signatures=with_signatures)
        sheets_created.append(sheet_title)

    return sheets_created


def generate_planilla_xlsx(
    *,
    event_name: str,
    event_date: datetime | None,
    event_location: str,
    operators_by_coordinator: dict[str, list[dict]] | None = None,
    operators: list[dict] | None = None,
    group_by: str = "coordinator",
    sort_by: str = "lastname",
    with_signatures: bool = False,
) -> bytes:
    """Genera el Excel de la planilla de pago.

    Modos soportados (combinaciones de ``group_by`` y ``sort_by``):

    ================== ============= =============
    group_by           sort_by       Descripción
    ================== ============= =============
    ``"coordinator"``  ``"lastname"``  Hoja por coordinador, ordenado por apellido (default)
    ``"coordinator"``  ``"document"``  Hoja por coordinador, ordenado por cédula
    ``"none"``         ``"lastname"``  Lista única, ordenada por apellido
    ``"none"``         ``"document"``  Lista única, ordenada por cédula
    ================== ============= =============

    Args:
        event_name: Nombre del evento.
        event_date: Fecha/hora de inicio del evento.
        event_location: Lugar del evento.
        operators_by_coordinator: (Retrocompatibilidad) Dict
            {coordinator_name: [op_data, ...]}. Si se pasa, se usa como
            fuente de datos (ignora ``operators`` y ``group_by`` queda
            forzado a ``"coordinator"``).
        operators: Lista plana de op_data. Necesario cuando
            ``group_by="none"``. Cada op_data tiene las claves:
            full_name, document_number, address, phone,
            coordinator_name, jacket_number, cap_number.
        group_by: ``"coordinator"``, ``"role"``, ``"coordinator_role"`` o ``"none"``.
        sort_by: ``"lastname"`` o ``"document"``.

    Returns:
        bytes con el contenido del archivo .xlsx listo para devolver como
        StreamingResponse.
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {TEMPLATE_PATH}")

    # Cargar plantilla
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = template_wb[TEMPLATE_SHEET]

    # --- Retrocompatibilidad: si llega operators_by_coordinator, usarlo ---
    if operators_by_coordinator is not None:
        operators = []
        for coord_name, ops in operators_by_coordinator.items():
            for op in ops:
                # Asegurar que cada op lleve su coordinator_name
                op_copy = dict(op)
                op_copy.setdefault("coordinator_name", coord_name)
                operators.append(op_copy)
        group_by = "coordinator"

    # Caso edge: sin operadores → devolver plantilla con datos del evento
    if not operators:
        _fill_header(
            template_ws,
            coordinator="SIN COORDINADOR ASIGNADO",
            event_name=event_name,
            event_date=event_date,
            event_location=event_location,
        )
        buf = io.BytesIO()
        template_wb.save(buf)
        return buf.getvalue()

    sheets_created: list[str] = []

    if group_by == "coordinator":
        # Agrupar operadores por coordinator_name (preservando el orden
        # de aparición para títulos estables).
        groups: dict[str, list[dict]] = {}
        for op in operators:
            coord_name = op.get("coordinator_name") or "SIN COORDINADOR"
            groups.setdefault(coord_name, []).append(op)

        for coord_name, ops in groups.items():
            sheets_created.extend(
                _render_pages(
                    template_wb=template_wb,
                    template_ws=template_ws,
                    event_name=event_name,
                    event_date=event_date,
                    event_location=event_location,
                    sheet_label=coord_name,
                    operators=ops,
                    sort_by=sort_by,
                    with_signatures=with_signatures,
                )
            )
    elif group_by == "role":
        # Agrupar operadores por role_name (preservando el orden de
        # aparición para títulos estables).
        groups: dict[str, list[dict]] = {}
        for op in operators:
            role_name = op.get("role_name") or "OPERADOR"
            groups.setdefault(role_name, []).append(op)

        for role_name, ops in groups.items():
            sheets_created.extend(
                _render_pages(
                    template_wb=template_wb,
                    template_ws=template_ws,
                    event_name=event_name,
                    event_date=event_date,
                    event_location=event_location,
                    sheet_label=role_name,
                    operators=ops,
                    sort_by=sort_by,
                    with_signatures=with_signatures,
                )
            )
    elif group_by == "coordinator_role":
        # Agrupar primero por coordinador y luego por rol dentro de cada
        # coordinador. El título de la hoja es "COORDINADOR - ROL".
        groups: dict[str, list[dict]] = {}
        for op in operators:
            coord_name = op.get("coordinator_name") or "SIN COORDINADOR"
            role_name = op.get("role_name") or "OPERADOR"
            label = f"{coord_name} - {role_name}"
            groups.setdefault(label, []).append(op)

        for label, ops in groups.items():
            sheets_created.extend(
                _render_pages(
                    template_wb=template_wb,
                    template_ws=template_ws,
                    event_name=event_name,
                    event_date=event_date,
                    event_location=event_location,
                    sheet_label=label,
                    operators=ops,
                    sort_by=sort_by,
                    with_signatures=with_signatures,
                )
            )
    else:
        # group_by == "none": lista única titulada con el nombre del evento
        sheets_created.extend(
            _render_pages(
                template_wb=template_wb,
                template_ws=template_ws,
                event_name=event_name,
                event_date=event_date,
                event_location=event_location,
                sheet_label=event_name or "EVENTO",
                operators=operators,
                sort_by=sort_by,
                with_signatures=with_signatures,
            )
        )

    # Eliminar la hoja plantilla original (si ya creamos al menos una)
    if sheets_created:
        del template_wb[TEMPLATE_SHEET]
        # La primera hoja queda como activa
        template_wb.active = 0

    # Serializar a bytes
    buf = io.BytesIO()
    template_wb.save(buf)
    logger.info(
        "Planilla generada: evento=%r, group_by=%s, sort_by=%s, hojas=%s",
        event_name,
        group_by,
        sort_by,
        sheets_created,
    )
    return buf.getvalue()
