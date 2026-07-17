"""Servicio para generar el Excel de novedades (incidencias) de un evento.

Genera un archivo ``.xlsx`` simple (sin plantilla externa) con una sola hoja
que lista todas las novedades registradas para un evento, ordenadas por fecha
ascendente. Pensado para imprimir y llevar registro físico de las novedades
operativas (llegadas tarde, incumplimientos, vetos, etc.).

Estructura de columnas:
    No | Fecha | Operador | Cédula | Tipo de Novedad | Descripción | Registró
"""
import io
import logging
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Zona horaria Bogotá (UTC-5, Colombia no usa horario de verano).
BOGOTA_TZ = timezone(timedelta(hours=-5))


def _to_bogota(dt) -> datetime | None:
    """Convierte un datetime a zona horaria Bogotá.

    - None → None
    - Naive datetime (sin tzinfo) → se asume UTC y se convierte a Bogotá
    - Datetime con tzinfo → se convierte a Bogotá

    Esto corrige el bug donde las horas salían 5h adelantadas (en UTC)
    en lugar de la hora local colombiana.
    """
    if dt is None:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return None
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(BOGOTA_TZ)

# Etiquetas legibles para los tipos de novedad.
TYPE_LABELS = {
    "doble_turno": "Doble turno",
    "llegada_tarde": "Llegada tarde",
    "salida_anticipada": "Salida anticipada",
    "incumplimiento": "Incumplimiento",
    "llamado_atencion": "Llamado de atención",
    "observacion": "Observación",
    "otro": "Otro",
    "veto": "Veto",
}

# Estilos reutilizables (colores marca A&C).
_HEADER_FILL = PatternFill(start_color="5D4224", end_color="5D4224", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFF2E4")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="5D4224")
_SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
_WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_novedades_xlsx(
    *,
    event_name: str,
    event_date: datetime | None,
    event_location: str,
    novedades: list[dict],
) -> bytes:
    """Genera el Excel de novedades para imprimir.

    Args:
        event_name: Nombre del evento.
        event_date: Fecha/hora de inicio del evento.
        event_location: Lugar del evento.
        novedades: Lista de dicts con claves:
            ``created_at`` (datetime), ``operator_name`` (str),
            ``operator_document`` (str), ``incident_type`` (str),
            ``description`` (str), ``recorder_name`` (str|None),
            ``is_veto`` (bool).

    Returns:
        bytes con el contenido del ``.xlsx``.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Novedades"

    # --- Encabezado informativo (filas 1-4) ---
    ws["A1"] = "NOVEDADES DEL EVENTO"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Evento: {event_name or '—'}"
    ws["A2"].font = _SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    fecha_str = ""
    bogota_date = _to_bogota(event_date)
    if bogota_date:
        fecha_str = bogota_date.strftime("%d/%m/%Y %H:%M")
    ws["A3"] = f"Fecha: {fecha_str or '—'}"
    ws["A3"].font = _SUBTITLE_FONT
    ws.merge_cells("A3:G3")

    ws["A4"] = f"Lugar: {event_location or '—'}"
    ws["A4"].font = _SUBTITLE_FONT
    ws.merge_cells("A4:G4")

    ws["A5"] = f"Total de novedades: {len(novedades)}"
    ws["A5"].font = Font(name="Calibri", size=10, bold=True, color="5D4224")
    ws.merge_cells("A5:G5")

    # --- Fila de encabezados de tabla (fila 7) ---
    HEADER_ROW = 7
    headers = ["No", "Fecha", "Operador", "Cédula", "Tipo de Novedad", "Descripción", "Registró"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER

    # --- Filas de datos (desde fila 8) ---
    data_row = HEADER_ROW + 1
    for idx, nov in enumerate(novedades, start=1):
        # Fecha formateada en zona horaria Bogotá (UTC-5).
        # La BD guarda created_at en UTC; sin esta conversión las horas
        # saldrían 5h adelantadas respecto a la hora local colombiana.
        bogota_created = _to_bogota(nov.get("created_at"))
        fecha_nov = bogota_created.strftime("%d/%m/%Y %H:%M") if bogota_created else "—"

        tipo = TYPE_LABELS.get(nov.get("incident_type", ""), nov.get("incident_type", "—"))
        if nov.get("is_veto"):
            tipo = f"🚫 {tipo}"

        row_values = [
            idx,
            fecha_nov,
            nov.get("operator_name") or "—",
            nov.get("operator_document") or "—",
            tipo,
            nov.get("description") or "—",
            nov.get("recorder_name") or "—",
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP_ALIGN if col_idx in (3, 6) else Alignment(vertical="top")
        data_row += 1

    # --- Anchos de columna (ajustados para impresión A4 horizontal) ---
    col_widths = {
        "A": 6,   # No
        "B": 16,  # Fecha
        "C": 32,  # Operador
        "D": 16,  # Cédula
        "E": 22,  # Tipo
        "F": 50,  # Descripción
        "G": 22,  # Registró
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Configuración de impresión ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    # Repetir fila de encabezados en cada página impresa.
    ws.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

    # Fijar el freeze pane justo debajo de los encabezados de tabla.
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # --- Serializar a bytes ---
    buf = io.BytesIO()
    wb.save(buf)
    logger.info(
        "Novedades Excel generado: evento=%r, total_novedades=%d",
        event_name,
        len(novedades),
    )
    return buf.getvalue()