"""Importación masiva de operadores a un evento desde Excel (.xlsx).

Servicio principal que:
  1. Lee un Excel con 16 columnas oficiales.
  2. Crea usuarios (user_type=operator) + perfil Operator para los nuevos.
  3. Asigna (nuevos y existentes) al evento con status='confirmed'.
  4. Resuelve el coordinador contra EventCoordinatorQuota del evento.
  5. Devuelve un ImportSummary con métricas y detalle fila por fila.

Decisiones del producto:
  - Estado de asignación: 'confirmed'.
  - Contraseña: NumeroDocumento + TipoDocumento (ej: 1027522598CC), sin acentos.
  - Email auto-generado: {documento}@operador.temp.
  - Coordinador: se resuelve por nombre contra cupos del evento; fallback texto libre.
  - Manejo tolerante: una fila con error NO detiene el resto.
  - Pre-carga batch de catálogos (evita N+1).
  - Commit único al final.
"""
import io
import re
import time
import unicodedata
import uuid
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.events import ImportRowResult, ImportSummary
from app.services.auth import hash_password


# ──────────────────────────────────────────────
#  Definición de columnas del Excel
# ──────────────────────────────────────────────

class ExcelColumn(dict):
    """Dict tipado para una columna del Excel."""
    pass


# Mapeo de columnas del Excel a claves internas.
# Los encabezados se comparan en UPPER + sin acentos para matching flexible.
EXPECTED_COLUMNS = [
    {"key": "primer_nombre",      "label": "PRIMER NOMBRE"},
    {"key": "segundo_nombre",     "label": "SEGUNDO NOMBRE"},
    {"key": "primer_apellido",    "label": "PRIMER APELLIDO"},
    {"key": "segundo_apellido",   "label": "SEGUNDO APELLIDO"},
    {"key": "document_type",      "label": "TIPO DE DOCUMENTO"},
    {"key": "document_number",    "label": "NUMERO CEDULA"},            # ← obligatoria
    {"key": "birth_date",         "label": "FECHA DE NACIMIENTO"},
    {"key": "role_name",          "label": "ROL ASIGANDO"},
    {"key": "gender",             "label": "GENERO"},
    {"key": "eps_name",           "label": "EPS"},
    {"key": "pension_fund_name",  "label": "PENSION"},
    {"key": "address",            "label": "DIRECCION DE VIVIENDA"},
    {"key": "phone",              "label": "NUMERO DE CELULAR"},
    {"key": "emergency_name",     "label": "NOMBRE CONTACTO EN CASO DE EMERGENCIA"},
    {"key": "emergency_phone",    "label": "TELEFONO CONTACTO EN CASO DE EMERGENCIA"},
    {"key": "coordinator_name",   "label": "COORDINADOR QUE LO PROGRAMA"},
]

# Columna obligatoria (clave interna).
REQUIRED_KEYS = {"document_number"}


# ──────────────────────────────────────────────
#  Helpers de normalización (portados del script CLI)
# ──────────────────────────────────────────────

def _strip_accents(s: str) -> str:
    """Quita acentos/tildes de un string."""
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm_header(s: str) -> str:
    """Normaliza un encabezado para matching flexible: UPPER + sin acentos + colapsar espacios."""
    if not s:
        return ""
    s = _strip_accents(str(s)).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _normalize_eps(name: str) -> str:
    """Normaliza un nombre de EPS para matching fuzzy."""
    if not name:
        return ""
    s = _strip_accents(name).upper().strip()
    for term in [
        "E.P.S.", "E.P.S", "EPS", "EPSS", "S.A.", "S.A", "SA",
        "LTDA.", "LTDA", "S.A.S.", "S.A.S", "SAS",
        "DE SALUD", "ENTIDAD PROMOTORA DE SALUD",
        "DIR.", "DIRECCION", "GENERAL", "FUERZAS", "MILITAR", "MILITARES",
        ".", ",", "-", "_",
    ]:
        s = s.replace(term, " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_phone(phone: str) -> str:
    """Limpia un teléfono: deja solo dígitos."""
    if not phone:
        return ""
    digits = re.sub(r"\D", "", str(phone))
    return digits[:20] if digits else ""


def _parse_date(date_val) -> Optional[datetime.date]:
    """Convierte varios formatos de fecha → date. Retorna None si falla."""
    if not date_val:
        return None
    # Si ya es un objeto date/datetime (openpyxl puede devolverlo)
    if hasattr(date_val, "year") and hasattr(date_val, "month"):
        return date_val if not hasattr(date_val, "hour") else date_val.date()
    s = str(date_val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _build_password(document_number: str, document_type: str) -> str:
    """Genera la contraseña: NumeroDocumento + TipoDocumento (sin acentos).

    Ejemplo: 1027522598 + CC = 1027522598CC
    """
    doc = _strip_accents(str(document_number or "")).strip()
    tipo = _strip_accents(str(document_type or "")).strip().upper()
    return f"{doc}{tipo}"


def _normalize_doc_type(raw: str) -> str:
    """Normaliza el tipo de documento a CC, CE, TI, etc."""
    if not raw:
        return "CC"
    s = _strip_accents(str(raw)).upper().strip()
    mapping = {
        "CEDULA DE CIUDADANIA": "CC",
        "CEDULA CIUDADANIA": "CC",
        "C.C.": "CC",
        "CC": "CC",
        "CEDULA DE EXTRANJERIA": "CE",
        "CEDULA EXTRANJERIA": "CE",
        "C.E.": "CE",
        "CE": "CE",
        "TARJETA DE IDENTIDAD": "TI",
        "TARJETA IDENTIDAD": "TI",
        "T.I.": "TI",
        "TI": "TI",
        "PASAPORTE": "PA",
        "PA": "PA",
    }
    return mapping.get(s, s[:10] if s else "CC")


def _compose_name(primer_nombre, segundo_nombre, primer_apellido, segundo_apellido):
    """Compone first_name y last_name desde las 4 columnas del Excel."""
    first_parts = [p for p in [primer_nombre, segundo_nombre] if p and str(p).strip()]
    last_parts = [p for p in [primer_apellido, segundo_apellido] if p and str(p).strip()]
    first_name = " ".join(str(p).strip() for p in first_parts) if first_parts else ""
    last_name = " ".join(str(p).strip() for p in last_parts) if last_parts else ""
    return first_name, last_name


# ──────────────────────────────────────────────
#  Matching de catálogos (roles, EPS, pensiones)
# ──────────────────────────────────────────────

def _match_role(role_name: str, role_map: dict) -> Optional[uuid.UUID]:
    """Matching fuzzy del rol. role_map = {norm_name: role_id}."""
    if not role_name:
        return None
    norm = _strip_accents(str(role_name)).upper().strip()
    norm_singular = re.sub(r"S$", "", norm)

    # Match exacto
    for key, rid in role_map.items():
        key_norm = _strip_accents(str(key)).upper().strip()
        if key_norm == norm or key_norm == norm_singular:
            return rid

    # Match parcial (contains)
    for key, rid in role_map.items():
        key_norm = _strip_accents(str(key)).upper().strip()
        if norm in key_norm or key_norm in norm:
            return rid
        if "BRIGADISTA" in norm and "BRIGADISTA" in key_norm:
            return rid
        if "OPERADOR" in norm and "LOGIST" in norm and "LOGIST" in key_norm:
            return rid

    return None


def _match_eps(eps_name: str, eps_list: list) -> Optional[uuid.UUID]:
    """Matching fuzzy de EPS. eps_list = [(id, norm_name, raw_name), ...]."""
    if not eps_name:
        return None
    norm = _normalize_eps(str(eps_name))

    # Match exacto normalizado
    for eid, ename_norm, _ in eps_list:
        if ename_norm == norm:
            return eid

    # Match parcial
    for eid, ename_norm, _ in eps_list:
        if not ename_norm or not norm:
            continue
        if ename_norm in norm or norm in ename_norm:
            return eid

    # Keywords
    keywords_map = {
        "SURA": "SURA", "SANITAS": "SANITAS", "NUEVA": "NUEVA",
        "COMPENSAR": "COMPENSAR", "FAMISANAR": "FAMISANAR",
        "SALUD TOTAL": "SALUD TOTAL", "CAPITAL SALUD": "CAPITAL SALUD",
        "SOS": "SERVICIO OCCIDENTAL",
    }
    norm_upper = norm.upper()
    for kw, eps_kw in keywords_map.items():
        if kw in norm_upper:
            for eid, ename_norm, _ in eps_list:
                if eps_kw.upper() in ename_norm.upper():
                    return eid
    return None


def _match_pension_fund(name: str, pf_list: list) -> Optional[uuid.UUID]:
    """Matching fuzzy de fondo de pensiones. pf_list = [(id, norm_name), ...]."""
    if not name:
        return None
    norm = _strip_accents(str(name)).upper().strip()
    for term in ["S.A.", "S.A", "SA", "LTDA.", "LTDA", "S.A.S.", "S.A.S", "SAS", ".", ","]:
        norm = norm.replace(term, " ")
    norm = re.sub(r"\s+", " ", norm).strip()

    # Match exacto
    for pid, pname_norm in pf_list:
        if pname_norm == norm:
            return pid

    # Match parcial
    for pid, pname_norm in pf_list:
        if not pname_norm or not norm:
            continue
        if pname_norm in norm or norm in pname_norm:
            return pid

    # Keywords comunes
    keywords_map = {
        "PORVENIR": "PORVENIR", "COLFONDOS": "COLFONDOS", "PROTECCION": "PROTECCION",
        "OLD MUTUAL": "OLD MUTUAL", "SKANDIA": "SKANDIA",
    }
    for kw in keywords_map:
        if kw in norm:
            for pid, pname_norm in pf_list:
                if kw in pname_norm:
                    return pid
    return None


# ──────────────────────────────────────────────
#  Generación del Excel plantilla
# ──────────────────────────────────────────────

def build_template() -> bytes:
    """Genera el .xlsx plantilla con encabezados oficiales (sin datos de ejemplo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Operadores"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Encabezados
    headers = [c["label"] for c in EXPECTED_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # (Sin fila de ejemplo: la plantilla se entrega vacía, solo con encabezados.)

    # Anchos de columna
    col_widths = [16, 16, 16, 16, 14, 16, 16, 22, 12, 22, 18, 28, 16, 28, 22, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Congelar primera fila
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
#  Lectura del Excel
# ──────────────────────────────────────────────

def _read_excel(file_bytes: bytes) -> list[dict]:
    """Lee el .xlsx y retorna lista de dicts (clave interna → valor).

    Hace matching flexible de encabezados (UPPER + sin acentos).
    Lanza ValueError con detalle si faltan columnas requeridas.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    # Leer encabezados reales del Excel (fila 1)
    raw_headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        raw_headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Mapear encabezados reales → clave interna (matching flexible)
    norm_to_key = {}
    for col in EXPECTED_COLUMNS:
        norm_to_key[_norm_header(col["label"])] = col["key"]

    header_map = {}  # {col_index: key}
    found_keys = set()
    for col_idx, raw_h in enumerate(raw_headers, 1):
        norm_h = _norm_header(raw_h)
        if norm_h in norm_to_key:
            key = norm_to_key[norm_h]
            if key not in found_keys:  # evitar duplicados
                header_map[col_idx] = key
                found_keys.add(key)

    # Validar columnas requeridas
    missing = REQUIRED_KEYS - found_keys
    if missing:
        labels = []
        for col in EXPECTED_COLUMNS:
            if col["key"] in missing:
                labels.append(col["label"])
        raise ValueError(
            f"Faltan columnas obligatorias en el Excel: {', '.join(labels)}"
        )

    # Leer filas de datos
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Saltar filas completamente vacías
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            continue
        row_dict = {}
        for col_idx, key in header_map.items():
            if col_idx <= len(row):
                val = row[col_idx - 1].value
                row_dict[key] = val if val is not None else None
            else:
                row_dict[key] = None
        rows.append(row_dict)

    wb.close()
    return rows


# ──────────────────────────────────────────────
#  Validación de una fila
# ──────────────────────────────────────────────

def _validate_row(row: dict, row_num: int) -> tuple[dict, list[str]]:
    """Valida una fila. Retorna (datos limpios, lista de errores)."""
    errors = []
    clean = {}

    # Documento obligatorio
    doc_raw = row.get("document_number")
    doc = str(doc_raw).strip() if doc_raw is not None else ""
    # Limpiar el documento: solo dígitos (algunos Excel vienen con formato numérico)
    if isinstance(doc_raw, float):
        doc = str(int(doc_raw))
    doc = re.sub(r"\D", "", doc)
    if not doc:
        errors.append("Número de documento vacío")
    clean["document_number"] = doc

    # Composición de nombres
    first_name, last_name = _compose_name(
        row.get("primer_nombre"),
        row.get("segundo_nombre"),
        row.get("primer_apellido"),
        row.get("segundo_apellido"),
    )
    clean["first_name"] = first_name or ""
    clean["last_name"] = last_name or "SIN APELLIDO"  # DB requiere non-null

    # Tipo de documento
    doc_type = _normalize_doc_type(row.get("document_type"))
    clean["document_type"] = doc_type

    # Fecha de nacimiento
    clean["birth_date"] = _parse_date(row.get("birth_date"))

    # Género
    gender_raw = row.get("gender")
    clean["gender"] = str(gender_raw).strip() if gender_raw else None

    # Rol
    clean["role_name"] = str(row.get("role_name") or "").strip() or None

    # EPS
    clean["eps_name"] = str(row.get("eps_name") or "").strip() or None

    # Pensión
    clean["pension_fund_name"] = str(row.get("pension_fund_name") or "").strip() or None

    # Dirección
    addr_raw = row.get("address")
    clean["address"] = str(addr_raw).strip() if addr_raw else None

    # Teléfono
    clean["phone"] = _normalize_phone(row.get("phone")) or None

    # Contacto de emergencia
    em_name = row.get("emergency_name")
    clean["emergency_name"] = str(em_name).strip() if em_name else None
    clean["emergency_phone"] = _normalize_phone(row.get("emergency_phone")) or None

    # Coordinador
    coord_raw = row.get("coordinator_name")
    clean["coordinator_name"] = str(coord_raw).strip() if coord_raw else None

    return clean, errors


# ──────────────────────────────────────────────
#  Pre-carga de catálogos (batch, evita N+1)
# ──────────────────────────────────────────────

async def _load_role_map(db: AsyncSession) -> dict:
    """Carga roles en un dict {norm_name: id}."""
    result = await db.execute(text("SELECT id, name FROM roles WHERE is_active = true"))
    role_map = {}
    for row in result:
        norm = _strip_accents(row.name).upper().strip()
        role_map[norm] = row.id
        role_map[row.name.strip().upper()] = row.id  # también sin normalizar
    return role_map


async def _load_eps_list(db: AsyncSession) -> list:
    """Carga EPS como [(id, norm_name, raw_name), ...]."""
    result = await db.execute(text("SELECT id, name FROM eps WHERE is_active = true"))
    return [(r.id, _normalize_eps(r.name), r.name) for r in result]


async def _load_pension_fund_list(db: AsyncSession) -> list:
    """Carga fondos de pensión como [(id, norm_name), ...]."""
    result = await db.execute(text("SELECT id, name FROM pension_fund"))
    return [(r.id, _strip_accents(r.name).upper().strip()) for r in result]


async def _load_coordinators(db: AsyncSession, event_id: uuid.UUID) -> list:
    """Carga coordinadores del evento desde EventCoordinatorQuota.

    Retorna [(operator_id, norm_name, display_name), ...].
    Incluye coordinadores legacy (sin operator_id) y del nuevo flujo (con FK).
    """
    result = await db.execute(
        text("""
            SELECT coordinator_operator_id, coordinator
            FROM event_coordinator_quotas
            WHERE event_id = :eid
        """),
        {"eid": str(event_id)},
    )
    coords = []
    for r in result:
        norm = _strip_accents(r.coordinator).upper().strip()
        op_id = str(r.coordinator_operator_id) if r.coordinator_operator_id else None
        coords.append((op_id, norm, r.coordinator))
    return coords


async def _ensure_coordinator_quotas(
    db: AsyncSession, event_id: uuid.UUID,
    excel_coords: dict, existing_coords: list,
) -> tuple[int, list]:
    """Crea quotas de coordinadores faltantes desde el Excel.

    Para cada coordinador del Excel que no tiene quota en el evento, crea una
    quota (legacy o con FK si se resuelve el operator_id). El cupo se calcula
    del conteo de operadores del Excel + un margen mínimo.

    Args:
        excel_coords: {norm_name: (display_name, count)} del Excel.
        existing_coords: lista de coordinadores ya con quota.

    Returns:
        (num_created, updated_coords_list) — coords actualizadas con las nuevas.
    """
    created = 0

    # Nombres normalizados ya existentes
    existing_norms = {c[1] for c in existing_coords}

    # Pre-cargar operadores de la BD para intentar resolver coordinator_operator_id
    # por nombre (los coordinadores suelen ser operadores del sistema).
    all_ops = await db.execute(text("""
        SELECT o.id, u.first_name, u.last_name, u.document_number
        FROM operators o
        JOIN users u ON o.user_id = u.id
        WHERE o.is_active = true
    """))
    ops_by_norm = {}
    for r in all_ops:
        full = f"{r.first_name or ''} {r.last_name or ''}".strip()
        norm = _strip_accents(full).upper().strip()
        ops_by_norm[norm] = str(r.id)
        # También indexar por primer nombre para matching parcial
        first_token = norm.split()[0] if norm.split() else ""
        if first_token and first_token not in ops_by_norm:
            ops_by_norm[first_token] = str(r.id)

    for cnorm, (display, count) in excel_coords.items():
        # ¿Ya existe esta quota? (match exacto o parcial)
        already = any(
            cnorm == ex_norm or cnorm in ex_norm or ex_norm in cnorm
            for ex_norm in existing_norms
        )
        if already:
            continue

        # Intentar resolver coordinator_operator_id por nombre
        coord_op_id = ops_by_norm.get(cnorm)
        if not coord_op_id:
            # Intentar por primer token
            first_token = cnorm.split()[0] if cnorm.split() else ""
            coord_op_id = ops_by_norm.get(first_token)

        # Cupo: operadores del Excel + margen (mínimo 5)
        quota_val = max(count + 5, 5)

        # Display name en MAYÚSCULAS
        display_upper = _strip_accents(display).upper().strip()

        try:
            await db.execute(text("""
                INSERT INTO event_coordinator_quotas (
                    id, event_id, coordinator, coordinator_operator_id, quota
                ) VALUES (
                    gen_random_uuid(), :eid, :coord, :op_id, :quota
                )
                ON CONFLICT DO NOTHING
            """), {
                "eid": str(event_id),
                "coord": display_upper,
                "op_id": coord_op_id,
                "quota": quota_val,
            })
            created += 1
            # Agregar a la lista existente para que las filas posteriores lo encuentren
            existing_coords.append((coord_op_id, cnorm, display_upper))
            existing_norms.add(cnorm)
        except Exception:
            # Si falla (p. ej. constraint), no detener la importación
            pass

    return created, existing_coords


def _resolve_coordinator(coord_name: str, coords: list) -> tuple[Optional[uuid.UUID], Optional[str]]:
    """Busca un coordinador por nombre (fuzzy) en la lista de coordinadores del evento.

    Retorna (operator_id, display_name). Si no hay match → (None, coord_name original).
    """
    if not coord_name:
        return None, None
    norm = _strip_accents(str(coord_name)).upper().strip()

    # Match exacto normalizado
    for op_id, c_norm, display in coords:
        if c_norm == norm:
            return op_id, display

    # Match parcial (contains)
    for op_id, c_norm, display in coords:
        if not c_norm or not norm:
            continue
        if c_norm in norm or norm in c_norm:
            return op_id, display

    # Match por primer nombre (ej: "XIMENA" → "XIMENA RODRIGUEZ")
    for op_id, c_norm, display in coords:
        first_token = c_norm.split()[0] if c_norm.split() else ""
        if first_token and first_token == norm:
            return op_id, display

    # No match → texto libre
    return None, _strip_accents(str(coord_name)).upper().strip()


async def _load_existing_users_by_doc(db: AsyncSession, doc_numbers: list[str]) -> dict:
    """Carga user_id+operator_id existentes por document_number.

    Retorna {document_number: (user_id, operator_id)}.
    """
    if not doc_numbers:
        return {}
    # Query en lotes para evitar SQL demasiado grande
    result = {}
    batch_size = 500
    for i in range(0, len(doc_numbers), batch_size):
        batch = doc_numbers[i:i + batch_size]
        placeholders = ",".join(f":d{j}" for j in range(len(batch)))
        params = {f"d{j}": batch[j] for j in range(len(batch))}
        rows = await db.execute(text(f"""
            SELECT u.id AS user_id, u.document_number,
                   (SELECT o.id FROM operators o WHERE o.user_id = u.id) AS operator_id
            FROM users u
            WHERE u.document_number IN ({placeholders})
        """), params)
        for r in rows:
            result[str(r.document_number)] = (r.user_id, r.operator_id)
    return result


async def _load_assigned_operators(db: AsyncSession, event_id: uuid.UUID) -> set:
    """Carga operator_ids ya asignados al evento."""
    result = await db.execute(
        text("SELECT operator_id FROM event_assignments WHERE event_id = :eid"),
        {"eid": str(event_id)},
    )
    return {r.operator_id for r in result}


# ──────────────────────────────────────────────
#  Orquestador principal
# ──────────────────────────────────────────────

async def import_operators_from_excel(
    db: AsyncSession, file_bytes: bytes, event_id: uuid.UUID,
) -> ImportSummary:
    """Procesa el Excel completo y devuelve un ImportSummary.

    Flujo:
      1. Lee el Excel.
      2. Pre-carga catálogos (roles, EPS, pensiones, coordinadores, usuarios existentes).
      3. Por cada fila: valida, crea/asigna, registra resultado.
      4. Commit único.
    """
    t0 = time.time()
    rows_result: list[ImportRowResult] = []
    created = 0
    existing = 0
    already_assigned = 0
    errors = 0

    # --- 1. Leer Excel ---
    try:
        raw_rows = _read_excel(file_bytes)
    except ValueError as exc:
        return ImportSummary(
            total_rows=0, created=0, existing=0, already_assigned=0,
            assigned=0, errors=1, duration_seconds=0.0,
            rows=[ImportRowResult(
                row=0, status="error",
                message=f"Error leyendo Excel: {exc}",
            )],
        )

    if not raw_rows:
        return ImportSummary(
            total_rows=0, created=0, existing=0, already_assigned=0,
            assigned=0, errors=0, duration_seconds=0.0, rows=[],
        )

    # --- 2. Pre-carga batch ---
    role_map = await _load_role_map(db)
    eps_list = await _load_eps_list(db)
    pf_list = await _load_pension_fund_list(db)
    coords = await _load_coordinators(db, event_id)

    # Documentos del Excel (para batch query de existentes)
    all_docs = []
    for raw in raw_rows:
        doc_val = raw.get("document_number")
        if doc_val is not None:
            d = str(doc_val).strip()
            if isinstance(doc_val, float):
                d = str(int(doc_val))
            d = re.sub(r"\D", "", d)
            if d:
                all_docs.append(d)

    existing_users = await _load_existing_users_by_doc(db, list(set(all_docs)))
    assigned_ops = await _load_assigned_operators(db, event_id)

    # --- 2b. Auto-crear quotas de coordinadores faltantes ---
    # Recolectar coordinadores únicos del Excel con su conteo de operadores.
    excel_coords: dict[str, tuple[str, int]] = {}  # {norm_name: (display, count)}
    for raw in raw_rows:
        coord_val = raw.get("coordinator_name")
        if not coord_val or not str(coord_val).strip():
            continue
        display = _strip_accents(str(coord_val)).upper().strip()
        cnorm = display
        if cnorm not in excel_coords:
            excel_coords[cnorm] = (display, 0)
        # Incrementar el conteo de operadores para este coordinador
        prev_display, prev_count = excel_coords[cnorm]
        excel_coords[cnorm] = (prev_display, prev_count + 1)

    # Crear quotas para coordinadores del Excel que no existen en el evento
    created_quotas, coords = await _ensure_coordinator_quotas(
        db, event_id, excel_coords, coords,
    )

    # Detección de duplicados dentro del Excel
    seen_docs: set[str] = set()

    now_utc = datetime.now(timezone.utc)

    # --- 3. Procesar filas ---
    for idx, raw in enumerate(raw_rows, 1):
        clean, val_errors = _validate_row(raw, idx)
        doc = clean["document_number"]
        full_name = f"{clean['first_name']} {clean['last_name']}".strip()

        # Error de validación
        if val_errors:
            errors += 1
            rows_result.append(ImportRowResult(
                row=idx, document_number=doc or None, full_name=full_name or None,
                status="error", message="; ".join(val_errors),
            ))
            continue

        # Duplicado dentro del mismo Excel
        if doc in seen_docs:
            errors += 1
            rows_result.append(ImportRowResult(
                row=idx, document_number=doc, full_name=full_name,
                status="error",
                message=f"Documento duplicado dentro del Excel: {doc}",
            ))
            continue
        seen_docs.add(doc)

        warnings = []

        # --- Resolver catálogos ---
        role_id = _match_role(clean["role_name"], role_map) if clean["role_name"] else None
        if clean["role_name"] and not role_id:
            warnings.append(f"Rol '{clean['role_name']}' no encontrado")

        eps_id = _match_eps(clean["eps_name"], eps_list) if clean["eps_name"] else None
        if clean["eps_name"] and not eps_id:
            warnings.append(f"EPS '{clean['eps_name']}' no encontrada")

        pf_id = _match_pension_fund(clean["pension_fund_name"], pf_list) if clean["pension_fund_name"] else None
        if clean["pension_fund_name"] and not pf_id:
            warnings.append(f"Pensión '{clean['pension_fund_name']}' no encontrada")

        coord_op_id, coord_display = _resolve_coordinator(clean["coordinator_name"], coords)

        # --- Caso 1: operador ya existe en BD ---
        if doc in existing_users:
            user_id, operator_id = existing_users[doc]
            if operator_id and operator_id in assigned_ops:
                # Ya asignado a este evento
                already_assigned += 1
                rows_result.append(ImportRowResult(
                    row=idx, document_number=doc, full_name=full_name,
                    status="already_assigned",
                    message="Ya estaba asignado a este evento",
                    operator_id=str(operator_id) if operator_id else None,
                    warnings=warnings,
                ))
                continue

            # Asignar existente al evento
            if operator_id:
                await _create_assignment(
                    db, event_id, operator_id, role_id,
                    coord_op_id, coord_display,
                )
                assigned_ops.add(operator_id)
                existing += 1
                rows_result.append(ImportRowResult(
                    row=idx, document_number=doc, full_name=full_name,
                    status="existing",
                    message="Operador existente asignado al evento",
                    operator_id=str(operator_id),
                    warnings=warnings,
                ))
                continue
            else:
                # Existe el user pero no el operator profile → crear profile
                operator_id = await _create_operator_profile(
                    db, user_id, eps_id, pf_id, clean, role_id,
                )
                existing_users[doc] = (user_id, operator_id)

        # --- Caso 2: operador nuevo ---
        else:
            # Generar email y contraseña
            email = f"{doc}@operador.temp"
            password = _build_password(doc, clean["document_type"])

            try:
                user_id = await _create_user(db, email, password, clean, role_id)
            except Exception as exc:
                errors += 1
                rows_result.append(ImportRowResult(
                    row=idx, document_number=doc, full_name=full_name,
                    status="error", message=f"Error creando usuario: {exc}",
                ))
                continue

            try:
                operator_id = await _create_operator_profile(
                    db, user_id, eps_id, pf_id, clean, role_id,
                )
            except Exception as exc:
                errors += 1
                rows_result.append(ImportRowResult(
                    row=idx, document_number=doc, full_name=full_name,
                    status="error", message=f"Error creando perfil: {exc}",
                ))
                continue

            existing_users[doc] = (user_id, operator_id)

        # --- Asignar al evento ---
        if operator_id and operator_id not in assigned_ops:
            await _create_assignment(
                db, event_id, operator_id, role_id,
                coord_op_id, coord_display,
            )
            assigned_ops.add(operator_id)
            created += 1
            rows_result.append(ImportRowResult(
                row=idx, document_number=doc, full_name=full_name,
                status="created",
                message="Operador creado y asignado al evento",
                operator_id=str(operator_id),
                warnings=warnings,
            ))

    # --- 4. Commit único ---
    await db.commit()

    elapsed = round(time.time() - t0, 2)

    return ImportSummary(
        total_rows=len(raw_rows),
        created=created,
        existing=existing,
        already_assigned=already_assigned,
        assigned=created + existing,
        errors=errors,
        duration_seconds=elapsed,
        rows=rows_result,
    )


# ──────────────────────────────────────────────
#  Funciones de creación (SQL crudo, como el script CLI)
# ──────────────────────────────────────────────

async def _create_user(db, email, password, clean, role_id) -> uuid.UUID:
    """Crea un User y retorna su id."""
    pw_hash = hash_password(password)
    result = await db.execute(text("""
        INSERT INTO users (
            id, email, password_hash, first_name, last_name,
            phone, document_type, document_number, user_type,
            role_id, is_verified, is_approved, is_active
        ) VALUES (
            gen_random_uuid(), :email, :pw, :first_name, :last_name,
            :phone, :doc_type, :doc, 'operator',
            :role_id, true, true, true
        )
        RETURNING id
    """), {
        "email": email, "pw": pw_hash,
        "first_name": clean["first_name"][:100],
        "last_name": clean["last_name"][:100],
        "phone": clean.get("phone"),
        "doc_type": clean["document_type"],
        "doc": clean["document_number"],
        "role_id": role_id,
    })
    return result.scalar()


async def _create_operator_profile(db, user_id, eps_id, pf_id, clean, role_id) -> uuid.UUID:
    """Crea un perfil Operator y retorna su id."""
    import json
    experience_roles_json = json.dumps([str(role_id)]) if role_id else None

    result = await db.execute(text("""
        INSERT INTO operators (
            id, user_id, eps_id, pension_fund_id, birth_date, gender, address,
            emergency_contact_name, emergency_contact_phone,
            whatsapp, background_check_status, total_events, is_active,
            experience_roles, has_protocol_experience, event_size_experience
        ) VALUES (
            gen_random_uuid(), :user_id, :eps_id, :pf_id, :birth_date, :gender, :address,
            :emergency_name, :emergency_phone,
            :whatsapp, 'pending', 0, true,
            :experience_roles, true, '100'
        )
        RETURNING id
    """), {
        "user_id": user_id,
        "eps_id": eps_id,
        "pf_id": pf_id,
        "birth_date": clean.get("birth_date"),
        "gender": clean.get("gender"),
        "address": clean.get("address"),
        "emergency_name": clean.get("emergency_name"),
        "emergency_phone": clean.get("emergency_phone"),
        "whatsapp": clean.get("phone"),
        "experience_roles": experience_roles_json,
    })
    return result.scalar()


async def _create_assignment(db, event_id, operator_id, role_id, coord_op_id, coord_display):
    """Crea un EventAssignment con status='confirmed' y datos del coordinador."""
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.execute(text("""
        INSERT INTO event_assignments (
            id, event_id, operator_id, role_id, status,
            confirmed_at, is_active, reminder_sent,
            programmed_by, admitted_by,
            programmed_by_operator_id, admitted_by_operator_id
        ) VALUES (
            gen_random_uuid(), :eid, :oid, :rid, 'confirmed',
            NOW(), true, false,
            :coord_display, :coord_display,
            :coord_op_id, :coord_op_id
        )
        ON CONFLICT DO NOTHING
    """), {
        "eid": str(event_id),
        "oid": str(operator_id),
        "rid": role_id,
        "coord_display": coord_display,
        "coord_op_id": str(coord_op_id) if coord_op_id else None,
    })