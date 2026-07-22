import pytest
import uuid
from io import BytesIO
import openpyxl
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timezone

from app.models.events import Event, EventAssignment
from app.models.operators import Operator
from app.models.users import User
from app.models.roles import Role
from app.services.auth import hash_password
from app.services.planilla_excel import generate_planilla_xlsx


def test_generate_planilla_xlsx_structure():
    event_name = "Boda de Prueba"
    event_date = datetime(2026, 6, 26, 18, 0, 0)
    event_location = "Salon Elegant"
    
    # 25 operators under one coordinator to test pagination (> 20)
    ops = []
    for i in range(25):
        ops.append({
            "full_name": f"Operador_{i} Apellido_{i}",
            "document_number": f"Doc_{i}",
            "address": f"Calle {i}",
            "phone": f"300{i:07d}",
            "coordinator_name": "Coordinador A",
            "jacket_number": f"J_{i}",
            "cap_number": f"C_{i}"
        })
        
    data = {
        "Coordinador A": ops
    }
    
    xlsx_bytes = generate_planilla_xlsx(
        event_name=event_name,
        event_date=event_date,
        event_location=event_location,
        operators_by_coordinator=data
    )
    
    assert isinstance(xlsx_bytes, bytes)
    
    # Load with openpyxl to inspect structure
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    
    # Since there are 25 operators under "Coordinador A", it should create 2 sheets:
    # "Coordinador A (1)" and "Coordinador A (2)"
    sheet_names = wb.sheetnames
    assert "Coordinador A (1)" in sheet_names
    assert "Coordinador A (2)" in sheet_names
    assert "Planilla" not in sheet_names # Original template sheet should be deleted
    
    # Check headers in sheet 1
    ws1 = wb["Coordinador A (1)"]
    assert ws1["D5"].value == "Coordinador A"
    assert ws1["D6"].value == event_name
    assert ws1["D7"].value == "26/06/2026"
    assert ws1["K7"].value == event_location
    
    # Check that first operator name is split
    assert ws1["C9"].value == "Operador_0"
    assert ws1["D9"].value == "Apellido_0"
    assert ws1["E9"].value == "Doc_0"
    assert ws1["F9"].value == "Calle 0"
    assert ws1["G9"].value == "3000000000"
    assert ws1["H9"].value == "Coordinador A"
    assert ws1["I9"].value == "J_0"
    assert ws1["J9"].value == "C_0"
    
    # Check that 20th operator is in sheet 1 (row 28)
    assert ws1["C28"].value == "Operador_19"
    
    # Check sheet 2 has the remaining 5 operators
    ws2 = wb["Coordinador A (2)"]
    assert ws2["C9"].value == "Operador_20"
    assert ws2["C13"].value == "Operador_24"
    assert ws2["C14"].value is None # row 14 (6th index in sheet 2) should be empty


@pytest.fixture
async def setup_payroll_event(db: AsyncSession):
    # Create Event
    event = Event(
        id=uuid.uuid4(),
        name="Evento Corporativo",
        start_date=datetime(2026, 7, 10, 8, 0, 0, tzinfo=timezone.utc),
        end_date=datetime(2026, 7, 10, 18, 0, 0, tzinfo=timezone.utc),
        location="Centro Convenciones",
        status="active"
    )
    db.add(event)
    
    # Create Coordinators & Roles
    role_coord = Role(
        id=uuid.uuid4(),
        name="Coordinador General",
        area="General",
        hierarchy_level=1,
        is_event_only=True
    )
    role_operator = Role(
        id=uuid.uuid4(),
        name="Logistico",
        area="Logistica",
        hierarchy_level=3,
        is_event_only=False
    )
    db.add(role_coord)
    db.add(role_operator)
    await db.flush()
    
    # Create Users (Admin, Coordinator, Operator)
    user_coord = User(
        id=uuid.uuid4(),
        email="coord@ayceventos.com",
        password_hash=hash_password("password"),
        first_name="Carlos",
        last_name="Coordinador",
        user_type="admin",
        document_number="77777",
        is_verified=True,
        is_approved=True
    )
    user_op1 = User(
        id=uuid.uuid4(),
        email="op1@test.com",
        password_hash=hash_password("password"),
        first_name="Juan",
        last_name="Perez",
        user_type="operator",
        document_number="11111",
        is_verified=True,
        is_approved=True
    )
    user_op2 = User(
        id=uuid.uuid4(),
        email="op2@test.com",
        password_hash=hash_password("password"),
        first_name="Maria",
        last_name="Gomez",
        user_type="operator",
        document_number="22222",
        is_verified=True,
        is_approved=True
    )
    db.add(user_coord)
    db.add(user_op1)
    db.add(user_op2)
    await db.flush()
    
    # Operator profiles
    op_profile1 = Operator(user_id=user_op1.id, city="Bogota", address="Calle 12", phone="3001111")
    op_profile2 = Operator(user_id=user_op2.id, city="Bogota", address="Calle 13", phone="3002222")
    op_coord_profile = Operator(user_id=user_coord.id, city="Bogota")
    db.add(op_profile1)
    db.add(op_profile2)
    db.add(op_coord_profile)
    await db.flush()
    
    # Assignments
    # Coordinator assignment (hierarchy level 1)
    assign_coord = EventAssignment(
        event_id=event.id,
        operator_id=op_coord_profile.id,
        role_id=role_coord.id,
        status="checked_in"
    )
    # Operator 1: checked_in -> should be included in payroll spreadsheet
    assign_op1 = EventAssignment(
        event_id=event.id,
        operator_id=op_profile1.id,
        role_id=role_operator.id,
        status="checked_in",
        jacket_number="J-10",
        cap_number="C-05"
    )
    # Operator 2: confirmed (not checked_in) -> should NOT be included in payroll spreadsheet
    assign_op2 = EventAssignment(
        event_id=event.id,
        operator_id=op_profile2.id,
        role_id=role_operator.id,
        status="confirmed"
    )
    
    db.add(assign_coord)
    db.add(assign_op1)
    db.add(assign_op2)
    
    await db.commit()
    return event, user_coord


@pytest.mark.asyncio
async def test_download_planilla_coordinador_endpoint(client: AsyncClient, admin_token: str, setup_payroll_event):
    event, _ = setup_payroll_event
    response = await client.get(
        f"/api/payroll/events/{event.id}/planilla-coordinador",
        headers={"Authorization": f"Bearer {admin_token}"}
    )
    assert response.status_code == 200
    assert response.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Content-Disposition" in response.headers
    assert "attachment; filename=" in response.headers["Content-Disposition"]
    
    # Load content into openpyxl
    wb = openpyxl.load_workbook(BytesIO(response.content))
    # It should have a sheet named after the Coordinator General "Carlos Coordinador"
    assert "Carlos Coordinador" in wb.sheetnames
    ws = wb["Carlos Coordinador"]
    # Check operator 1 details are in the sheet
    assert ws["C9"].value == "Juan"
    assert ws["D9"].value == "Perez"
    assert ws["E9"].value == "11111"
    # Ensure operator 2 is not in the sheet since status was 'confirmed' and not 'checked_in'
    for row in range(9, 29):
        if ws.cell(row=row, column=5).value == "22222":
            pytest.fail("Operator 2 should not be in the spreadsheet")


@pytest.mark.asyncio
async def test_download_planilla_coordinador_forbidden(client: AsyncClient, operator_token: str, setup_payroll_event):
    event, _ = setup_payroll_event
    response = await client.get(
        f"/api/payroll/events/{event.id}/planilla-coordinador",
        headers={"Authorization": f"Bearer {operator_token}"}
    )
    assert response.status_code == 403
